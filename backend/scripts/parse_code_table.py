"""
수치지도 지형지물 표준코드 엑셀 → layer_codes.json 변환 스크립트

사용법:
  pip install openpyxl
  python parse_code_table.py <input.xlsx> [--output-dir ./data]

입력: 국가법령정보센터에서 다운로드한 "별표1 수치지도 지형지물 표준코드" 엑셀 파일
  - .xls인 경우 LibreOffice로 .xlsx 변환 후 사용
  - 변환: libreoffice --headless --convert-to xlsx input.xls

출력:
  - layer_codes.json (680개 항목, 전체 코드 테이블)
  - category_colors.json (8개 대분류 색상)
"""

import json
import argparse
from pathlib import Path
from openpyxl import load_workbook


DISPLAY_COLORS = {
    "A": "#FF6B6B",  # 교통
    "B": "#FFD32A",  # 건물
    "C": "#00CEC9",  # 시설
    "D": "#55EFC4",  # 식생
    "E": "#4D9FFF",  # 수계
    "F": "#E17055",  # 지형
    "G": "#B2BEC3",  # 경계
    "H": "#636E72",  # 주기
}

CATEGORY_NAMES = {
    "A": "교통", "B": "건물", "C": "시설", "D": "식생",
    "E": "수계", "F": "지형", "G": "경계", "H": "주기",
}


def parse_rgb(color_raw: str) -> str:
    if not color_raw or "R/G/B" not in str(color_raw):
        return "#888888"
    try:
        rgb_line = str(color_raw).replace("R/G/B\n", "").replace("R/G/B\\n", "").strip()
        parts = rgb_line.split("/")
        if len(parts) == 3:
            r, g, b = int(parts[0]), int(parts[1]), int(parts[2])
            return f"#{r:02x}{g:02x}{b:02x}"
    except (ValueError, IndexError):
        pass
    return "#888888"


def refine_display_color(letter: str, name: str, mid: str) -> str:
    base = DISPLAY_COLORS.get(letter, "#888888")
    if letter == "A":
        if "중심선" in name or "중심선" in mid:
            return "#FF9F9F"
        if "철도" in name or "철도" in mid:
            return "#A29BFE"
    elif letter == "E":
        if "중심선" in name or "중심선" in mid:
            return "#85C1FF"
    return base


def parse_excel(xlsx_path: str) -> tuple[dict, dict]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    layer_codes = {}
    current_major = ""
    current_mid = ""

    for row in ws.iter_rows(min_row=6, values_only=True):
        major, mid, name, code = row[2], row[3], row[4], row[5]
        struct_1000, struct_5000, struct_20 = row[12], row[6], row[15]
        color_1000, color_2500, color_5000 = row[14], row[11], row[8]

        if major: current_major = major
        if mid: current_mid = mid
        if not code or not name: continue

        first_letter = code[0] if code else ""
        color_raw = color_1000 or color_2500 or color_5000 or ""
        structure = str(struct_1000 or struct_5000 or struct_20 or "").strip()

        layer_codes[code] = {
            "code": code,
            "category_major": first_letter,
            "category_major_name": current_major,
            "category_mid": current_mid,
            "name": name,
            "official_color": parse_rgb(str(color_raw)),
            "default_color": refine_display_color(first_letter, name, current_mid),
            "linetype": "DASHED" if "중심선" in name else "Continuous",
            "structure": structure,
        }

    wb.close()

    category_colors = {
        letter: {"name": name, "color": DISPLAY_COLORS[letter]}
        for letter, name in CATEGORY_NAMES.items()
    }
    return layer_codes, category_colors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input", help="Path to .xlsx file")
    parser.add_argument("--output-dir", default=".", help="Output directory")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    layer_codes, category_colors = parse_excel(args.input)

    for name, data in [("layer_codes.json", layer_codes), ("category_colors.json", category_colors)]:
        with open(output_dir / name, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Generated {len(layer_codes)} entries")
    for letter in sorted(CATEGORY_NAMES.keys()):
        count = sum(1 for e in layer_codes.values() if e["category_major"] == letter)
        print(f"  {letter} {CATEGORY_NAMES[letter]}: {count}")


if __name__ == "__main__":
    main()
